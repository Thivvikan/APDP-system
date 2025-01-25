from abc import ABC, abstractmethod
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook, load_workbook
import pandas as pd
import matplotlib.pyplot as plt

# open close and template design pattern
class Entity(ABC):
    @staticmethod
    @abstractmethod
    def insert():
        pass

    @staticmethod
    @abstractmethod
    def update():
        pass

    @staticmethod
    @abstractmethod
    def delete():
        pass

    @staticmethod
    @abstractmethod
    def search():
        pass

# branch class
class Branch(Entity):
    def __init__(self, branch_id, name, location,contact_number):
        self.branch_id = branch_id
        self.name = name
        self.location = location
        self.contact_number=contact_number
        self.file = f"{branch_id}.xlsx"
        Sampath_food_cities.branches[branch_id] = {
            "name": self.name,
            "location": self.location,
            "contact_number":self.contact_number,
            "file": self.file
        }

    @staticmethod
    def insert():
        if not os.path.exists(branch_file):
            Sampath_food_cities.create_branches_file()

        work_book = load_workbook(branch_file)
        work_sheet = work_book.active

        # Generate Branch ID based on the current number of rows
        branch_id = (lambda row_number: f"Br{str(row_number).zfill(3)}")(work_sheet.max_row)

        # Take input for the new branch
        branch_name = input("Enter the branch name: ")
        location = input("Enter the location of the branch: ")
        contact_number=input("Enter the contact number for the branch: ")

        # Append new branch to the branch file
        work_sheet.append([branch_id, branch_name, location, contact_number])
        work_book.save(branch_file)

        print(f"Branch {branch_name} is saved successfully with the ID: {branch_id}")

        # Store branch info in the dictionary
        Sampath_food_cities.branches[branch_id] = {
            "name": branch_name,
            "location": location,
            "file": f"{branch_id}.xlsx"
        }
        print(f"Branch data for {branch_name} added to dictionary.")
        Sampath_food_cities.create_branch_management_file(branch_id)

    @staticmethod
    def update():
        if not os.path.exists(branch_file):
            print("Branch file does not exist. No branches to update.")
            return
        work_book = load_workbook(branch_file)
        work_sheet = work_book.active
        search_key = input("Enter Branch ID or Name to search: ").lower()
        for row in work_sheet.iter_rows(min_row=2, values_only=False):  # Skip header row
            branch_id, branch_name, location, contact_number = row
            if search_key == branch_id.value.lower() or search_key in branch_name.value.lower():
                print(f"Current details - Branch Name: {branch_name.value}, Location: {location.value}, Contact number: {contact_number.value}")
                # Update details
                new_name = input("Enter the new branch name (leave blank to keep current): ")
                new_location = input("Enter the new location (leave blank to keep current): ")
                new_contact_number=input("Enter the new contact number (leave blank to keep current): ")
                if new_name:
                    branch_name.value = new_name
                if new_location:
                    location.value = new_location
                if new_contact_number:
                    contact_number.value = new_contact_number
                work_book.save(branch_file)
                print(f"Branch {branch_id.value} updated successfully.")
                # Update dictionary
                Sampath_food_cities.branches[branch_id.value]["name"] = branch_name.value
                Sampath_food_cities.branches[branch_id.value]["location"] = location.value
                Sampath_food_cities.branches[branch_id.value]["contact_number"]=contact_number.value
                return
        print(f"No branch found for: {search_key}")

    @staticmethod
    def delete():
        if not os.path.exists(branch_file):
            print("Branch file does not exist. No branches to delete.")
            return
        work_book = load_workbook(branch_file)
        work_sheet = work_book.active
        search_key = input("Enter the Branch ID or Branch Name to delete: ").lower()
        # Track if any branch is found
        branch_found = False
        for row_index, row in enumerate(work_sheet.iter_rows(min_row=2, values_only=False), start=2):
            branch_id = row[0].value
            branch_name = row[1].value
            location = row[2].value
            contact_number = row[3].value
            if (branch_id and search_key == str(branch_id).lower()) or (
                    branch_name and search_key in branch_name.lower()):
                print(f"\nBranch Found:")
                print(f"Branch ID: {branch_id}")
                print(f"Branch Name: {branch_name}")
                print(f"Location: {location}")
                print(f"Contact Number: {contact_number}")
                # Confirm deletion
                confirmation = input(f"Are you sure you want to delete this branch? (yes/no): ").lower()
                if confirmation == "yes":
                    # Delete the row
                    work_sheet.delete_rows(row_index)
                    work_book.save(branch_file)
                    print(f"Branch '{branch_name}' (ID: {branch_id}) deleted successfully.")
                    # Remove from the dictionary if exists
                    if hasattr(Sampath_food_cities, "branches") and branch_id in Sampath_food_cities.branches:
                        Sampath_food_cities.branches.pop(branch_id, None)
                    # Delete the corresponding branch file
                    branch_file_to_delete = f"{branch_id}.xlsx"
                    if os.path.exists(branch_file_to_delete):
                        os.remove(branch_file_to_delete)
                        print(f"Branch file '{branch_file_to_delete}' deleted successfully.")
                    else:
                        print(f"No file found for branch ID {branch_id}.")
                    branch_found = True
                    break  # Exit the loop after deletion
        if not branch_found:
            print(f"No branch found with ID or Name: '{search_key}'.")

    @staticmethod
    def search():
        if not os.path.exists(branch_file):
            print("Branch file does not exist. No branches to search.")
            return

        work_book = load_workbook(branch_file)
        work_sheet = work_book.active

        search_query = input("Enter Branch ID or Name to search: ").lower()

        for row in work_sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            branch_id, branch_name, location, contact_number = row
            if search_query == branch_id.lower() or search_query in branch_name.lower():
                print(f"Branch Found - ID: {branch_id}, Name: {branch_name}, Location: {location}, Contact Number: {contact_number}")
                return

        print("No matching branch found.")

# product class
class Product(Entity):
    def __init__(self, product_id, name, unit_price):
        self.product_id = product_id
        self.name = name
        self.unit_price = unit_price

    @staticmethod
    def insert():
        """Insert a new product into the product file and update branch sales files."""
        # Check if the product file exists; if not, create it
        if not os.path.exists(product_file):
            Sampath_food_cities.create_Products_file()

        # Open the product file
        work_book = load_workbook(product_file)
        work_sheet = work_book.active

        # Generate product ID based on the current number of rows
        product_id = (lambda row_number: f"Pr{str(row_number).zfill(3)}")(work_sheet.max_row)

        # Get product details from user input
        product_name = input("Enter the product name: ")
        unit_price = float(input("Enter the unit price of the product: "))

        # Get the current date
        current_date = datetime.today().strftime('%d/%m/%Y')

        # Check if the date column exists in the product file
        headers = [cell.value for cell in work_sheet[1]]
        if current_date not in headers:
            # Add a new column for the date
            work_sheet.cell(row=1, column=len(headers) + 1, value=current_date)

        # Find the correct column index for the date
        date_column_index = headers.index(current_date) + 1 if current_date in headers else len(headers) + 1

        # Append the new product details to the product file
        row_data = [product_id, product_name, unit_price] + [""] * (date_column_index - 4) + [unit_price]
        work_sheet.append(row_data)

        # Save the product file
        work_book.save(product_file)
        print(f"Product {product_name} is saved successfully with the ID: {product_id}")

        # Update branch sales files with the new product
        Update_Product_info.update_in_branch_sales_files(product_id, product_name, unit_price)

    @staticmethod
    def update():
        # Prompt the user for the date
        update_date = input("Enter the date for price update (DD/MM/YYYY): ")
        # Check if the product file exists
        if not os.path.exists(product_file):
            print(f"Error: '{product_file}' does not exist. Cannot update prices.")
            return
        # Open the product file
        work_book = load_workbook(product_file)
        work_sheet = work_book.active
        # Add the date column if it does not exist
        if update_date not in [cell.value for cell in work_sheet[1]]:
            work_sheet.cell(row=1, column=work_sheet.max_column + 1).value = update_date
        # Display available products
        print("\nAvailable Products:")
        for row in work_sheet.iter_rows(min_row=2, max_row=work_sheet.max_row, values_only=True):
            print(f"Product ID: {row[0]}, Product Name: {row[1]}, Unit Price: {row[2]}")
        # Prompt the user for the product to update
        product_id = input("\nEnter the Product ID to update the price: ")
        # Find the product row
        product_row = None
        for row in work_sheet.iter_rows(min_row=2, max_row=work_sheet.max_row):
            if row[0].value == product_id:
                product_row = row
                break
        if not product_row:
            print(f"Error: Product ID '{product_id}' not found.")
            return
        # Prompt for the new price
        new_price = float(input(f"Enter the new price for {product_row[1].value}: "))
        # Update the unit price column and the new date column
        product_row[2].value = new_price  # Update unit price
        product_row[work_sheet.max_column - 1].value = new_price  # Update for the new date column
        # Save the product file
        work_book.save(product_file)
        print(f"Price updated successfully for {product_row[1].value} in '{product_file}'.")

        # Update price in all branch sales files
        Update_Product_info.update_branch_prices(product_id, new_price)

    @staticmethod
    def delete():
        """Delete a product from the product file and update branch sales files."""
        # Ensure the product file exists
        if not os.path.exists(product_file):
            print("Product file does not exist. Nothing to delete.")
            return

        # Open the product file
        work_book = load_workbook(product_file)
        work_sheet = work_book.active

        # Ask the user for the Product ID or Product Name
        search_key = input("Enter Product ID or Name to delete: ")

        # Find the product row
        row_to_delete = None
        for row_index, row in enumerate(work_sheet.iter_rows(min_row=2, max_row=work_sheet.max_row), start=2):
            product_id = row[0].value
            product_name = row[1].value
            if search_key == product_id or search_key.lower() == product_name.lower():
                row_to_delete = row_index
                break

        if not row_to_delete:
            print(f"No product found with ID or Name '{search_key}'.")
            return

        # Get product details for reference
        product_id = work_sheet.cell(row=row_to_delete, column=1).value
        product_name = work_sheet.cell(row=row_to_delete, column=2).value

        # Delete the row
        work_sheet.delete_rows(row_to_delete)
        work_book.save(product_file)
        print(f"Product '{product_name}' with ID '{product_id}' has been deleted successfully.")

        # Update branch sales files to remove the product
        Update_Product_info.remove_from_branch_sales_files(product_id)

    @staticmethod
    def search():
        """Search for a product in the product file."""
        # Ensure the product file exists
        if not os.path.exists(product_file):
            print("Product file does not exist. Please add products first.")
            return

        # Open the product file
        work_book = load_workbook(product_file)
        work_sheet = work_book.active

        # Ask the user for the Product ID or Product Name
        search_key = input("Enter Product ID or Name to search: ")

        # Search for the product in the product file
        for row in work_sheet.iter_rows(min_row=2, max_row=work_sheet.max_row):
            product_id = row[0].value
            product_name = row[1].value
            unit_price = row[2].value

            if search_key == product_id or search_key.lower() == product_name.lower():
                print(f"Product found: ID = {product_id}, Name = {product_name}, Unit Price = {unit_price}")
                return

        print(f"No product found with ID or Name '{search_key}'.")

# class purchase
class Purchase(Entity):
    @staticmethod
    def insert():
        # Ensure branches are available
        if not Sampath_food_cities.branches:
            print("No branches available. Please insert a branch first.")
            return
        # Ask for Branch ID or Branch Name
        search_key = input("Enter the Branch ID or Branch Name: ")
        branch_id = None
        # Search the branch in the dictionary
        for b_id, b_info in Sampath_food_cities.branches.items():
            if search_key == b_id or search_key.lower() == b_info["name"].lower():
                branch_id = b_id
                break
        if not branch_id:
            print(f"No branch found with ID or Name '{search_key}'.")
            return
        # Get the branch file
        branch_file = f"{branch_id}.xlsx"
        products_file = "products.xlsx"
        # Load the branch workbook
        if os.path.exists(branch_file):
            branch_workbook = load_workbook(branch_file)
            if "Monthly Purchases" in branch_workbook.sheetnames:
                purchase_sheet = branch_workbook["Monthly Purchases"]
            else:
                print(f"Error: 'Purchases' sheet not found in {branch_file}.")
                return
        else:
            print(f"Error: Branch file '{branch_file}' does not exist.")
            return
        # Ask for the month
        month = input("Enter the month for purchases (e.g., January 2024): ")
        # Check if the month column already exists, otherwise add it
        month_col_index = None
        for col_index, cell in enumerate(purchase_sheet[1], start=1):
            if cell.value == month:
                month_col_index = col_index
                break
        if not month_col_index:
            month_col_index = purchase_sheet.max_column + 1
            purchase_sheet.cell(row=1, column=month_col_index, value=month)
        # Start inserting purchases for products
        print("Enter purchases for products. Type 'done' to finish.")
        while True:
            # Ask for Product ID or Product Name
            product_id_or_name = input("Enter Product ID or Name (or type 'done' to finish): ")
            if product_id_or_name.lower() == "done":
                break
            # Find the product in the purchase file
            product_row_index = None
            for row_index, row in enumerate(purchase_sheet.iter_rows(min_row=2, max_row=purchase_sheet.max_row),
                                            start=2):
                if product_id_or_name == str(row[0].value) or product_id_or_name.lower() == str(row[1].value).lower():
                    product_row_index = row_index
                    break
            # If the product is not in the Purchases sheet, check in the products file
            if not product_row_index:
                products_workbook = load_workbook(products_file)
                products_sheet = products_workbook.active
                # Search for the product in the products file
                for row in products_sheet.iter_rows(min_row=2, max_row=products_sheet.max_row):
                    if product_id_or_name == str(row[0].value) or product_id_or_name.lower() == str(
                            row[1].value).lower():
                        product_id = row[0].value
                        product_name = row[1].value
                        # Add the product to the Purchases sheet
                        purchase_sheet.append([product_id, product_name])
                        product_row_index = purchase_sheet.max_row
                        print(f"Product '{product_name}' added to the Purchases sheet.")
                        break
                else:
                    print(f"No product found with ID or Name '{product_id_or_name}' in the products file.")
                    continue
            # Ask for the quantity
            try:
                quantity = int(
                    input(f"Enter quantity for {purchase_sheet.cell(row=product_row_index, column=2).value}: "))
            except ValueError:
                print("Invalid input for quantity. Please enter a valid number.")
                continue
            # Update the quantity in the correct cell
            purchase_sheet.cell(row=product_row_index, column=month_col_index, value=quantity)
        # Save the updated branch file
        branch_workbook.save(branch_file)
        print(
            f"Purchase details for branch '{Sampath_food_cities.branches[branch_id]['name']}' for {month} have been successfully updated.")

    @staticmethod
    def update():
        if not Sampath_food_cities.branches:
            print("No branches available. Please insert a branch first.")
            return

        search_key = input("Enter the Branch ID or Branch Name: ")
        branch_id = None

        # Search the branch in the dictionary
        for b_id, b_info in Sampath_food_cities.branches.items():
            if search_key == b_id or search_key.lower() == b_info["name"].lower():
                branch_id = b_id
                break

        if not branch_id:
            print(f"No branch found with ID or Name '{search_key}'.")
            return

        branch_file = f"{branch_id}.xlsx"
        if not os.path.exists(branch_file):
            print(f"Error: Branch file '{branch_file}' does not exist.")
            return

        branch_workbook = load_workbook(branch_file)
        if "Monthly Purchases" not in branch_workbook.sheetnames:
            print(f"Error: 'Monthly Purchases' sheet not found in {branch_file}.")
            return

        purchase_sheet = branch_workbook["Monthly Purchases"]
        month = input("Enter the month to update purchases (e.g., January 2024): ")

        # Locate the column for the month
        month_col_index = None
        for col_index, cell in enumerate(purchase_sheet[1], start=1):
            if cell.value == month:
                month_col_index = col_index
                break

        if not month_col_index:
            print(f"No data found for the month '{month}'. Please ensure purchases exist for this month.")
            return

        product_id_or_name = input("Enter Product ID or Name to update: ")
        for row in purchase_sheet.iter_rows(min_row=2, values_only=False):
            if product_id_or_name == str(row[0].value) or product_id_or_name.lower() == str(row[1].value).lower():
                try:
                    new_quantity = int(input(f"Enter new quantity for {row[1].value}: "))
                    row[month_col_index - 1].value = new_quantity
                    branch_workbook.save(branch_file)
                    print(f"Quantity updated for {row[1].value} in {month}.")
                    return
                except ValueError:
                    print("Invalid input for quantity. Please enter a valid number.")
                    return

        print(f"No product found with ID or Name '{product_id_or_name}' in branch purchases.")

    @staticmethod
    def delete():
        if not Sampath_food_cities.branches:
            print("No branches available. Please insert a branch first.")
            return

        search_key = input("Enter the Branch ID or Branch Name: ")
        branch_id = None

        for b_id, b_info in Sampath_food_cities.branches.items():
            if search_key == b_id or search_key.lower() == b_info["name"].lower():
                branch_id = b_id
                break

        if not branch_id:
            print(f"No branch found with ID or Name '{search_key}'.")
            return

        branch_file = f"{branch_id}.xlsx"
        if not os.path.exists(branch_file):
            print(f"Error: Branch file '{branch_file}' does not exist.")
            return

        branch_workbook = load_workbook(branch_file)
        if "Monthly Purchases" not in branch_workbook.sheetnames:
            print(f"Error: 'Monthly Purchases' sheet not found in {branch_file}.")
            return

        purchase_sheet = branch_workbook["Monthly Purchases"]
        product_id_or_name = input("Enter Product ID or Name to delete: ")

        for row_index, row in enumerate(purchase_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if product_id_or_name == str(row[0]) or product_id_or_name.lower() == str(row[1]).lower():
                purchase_sheet.delete_rows(row_index)
                branch_workbook.save(branch_file)
                print(f"Product {row[1]} has been deleted from the purchases.")
                return

        print(f"No product found with ID or Name '{product_id_or_name}' in branch purchases.")

    @staticmethod
    def search():
        if not Sampath_food_cities.branches:
            print("No branches available. Please insert a branch first.")
            return

        search_key = input("Enter the Branch ID or Branch Name: ")
        branch_id = None

        for b_id, b_info in Sampath_food_cities.branches.items():
            if search_key == b_id or search_key.lower() == b_info["name"].lower():
                branch_id = b_id
                break

        if not branch_id:
            print(f"No branch found with ID or Name '{search_key}'.")
            return

        branch_file = f"{branch_id}.xlsx"
        if not os.path.exists(branch_file):
            print(f"Error: Branch file '{branch_file}' does not exist.")
            return

        branch_workbook = load_workbook(branch_file)
        if "Monthly Purchases" not in branch_workbook.sheetnames:
            print(f"Error: 'Monthly Purchases' sheet not found in {branch_file}.")
            return

        purchase_sheet = branch_workbook["Monthly Purchases"]
        month = input("Enter the month to search (e.g., January 2024): ")

        # Locate the column for the month
        month_col_index = None
        for col_index, cell in enumerate(purchase_sheet[1], start=1):
            if cell.value == month:
                month_col_index = col_index
                break

        if not month_col_index:
            print(f"No data found for the month '{month}'.")
            return

        print(f"Purchases for the month '{month}':")
        for row in purchase_sheet.iter_rows(min_row=2, values_only=True):
            product_id, product_name, *data = row
            quantity = data[month_col_index - 3] if len(data) >= (month_col_index - 2) else 0
            print(f"Product ID: {product_id}, Product Name: {product_name}, Quantity: {quantity}")

# class update product info
class Update_Product_info:
    @staticmethod
    def update_in_branch_sales_files(product_id, product_name, unit_price):
        # Update all branch sales files with new product details
        for branch in Sampath_food_cities.branches.values():
            branch_file = branch['file']
            if os.path.exists(branch_file):
                branch_workbook = load_workbook(branch_file)
                # Update Daily Sales and Product Quantity sheets
                for sheet_name in ["Daily Sales in branch", "Product Quantity"]:
                    if sheet_name not in branch_workbook.sheetnames:
                        branch_workbook.create_sheet(sheet_name)
                    sheet = branch_workbook[sheet_name]
                    sheet.append([product_id, product_name, unit_price])
                branch_workbook.save(branch_file)
                print(f"Product {product_name} added to '{branch_file}'.")

    @staticmethod
    def update_branch_prices(product_id, new_price):
        # Check if the branch file exists
        for branch in Sampath_food_cities.branches.values():
            branch_file = branch['file']
            if os.path.exists(branch_file):
                branch_workbook = load_workbook(branch_file)
                # Find the product in the branch file and update its price
                for sheet_name in ["Daily Sales in branch", "Product Quantity"]:
                    if sheet_name in branch_workbook.sheetnames:
                        sheet = branch_workbook[sheet_name]
                        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                            if row[0].value == product_id:
                                row[2].value = new_price  # Update the unit price
                                break
                branch_workbook.save(branch_file)
                print(f"Price updated successfully for {product_id} in branch sales file '{branch_file}'.")

    @staticmethod
    def remove_from_branch_sales_files(product_id):
        """Remove a product from all branch sales files."""
        # Iterate through all branches in Sampath_food_cities
        for branch in Sampath_food_cities.branches.values():
            branch_file = branch['file']

            # Check if the branch file exists
            if os.path.exists(branch_file):
                branch_workbook = load_workbook(branch_file)

                # Iterate through the relevant sheets in the branch file
                for sheet_name in ["Daily Sales in branch", "Product Quantity"]:
                    if sheet_name in branch_workbook.sheetnames:
                        sheet = branch_workbook[sheet_name]

                        # Search for the product in the sheet
                        row_to_delete = None
                        for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
                            if row[0].value == product_id:  # Check if product ID matches
                                row_to_delete = row_index
                                break

                        # If the product is found, delete the row
                        if row_to_delete:
                            sheet.delete_rows(row_to_delete)
                            print(
                                f"Product ID {product_id} removed from '{sheet_name}' in branch file '{branch_file}'.")

                # Save changes to the branch file
                branch_workbook.save(branch_file)
            else:
                print(f"Branch file '{branch_file}' does not exist. Skipping this branch.")

        print(f"Product ID {product_id} removed from all branch sales files successfully.")

    @staticmethod
    def copy_products_to_branch_management_file(branch_id):
        # Define the product file and branch file names
        product_file = "Products.xlsx"
        branch_file = f"{branch_id}.xlsx"

        # Verify if the product file exists
        if not os.path.exists(product_file):
            print(f"Error: '{product_file}' does not exist. Cannot copy product data.")
            return

        # Verify if the branch file exists
        if not os.path.exists(branch_file):
            print(f"Error: Branch file '{branch_file}' does not exist.")
            return

        # Load the product file and its sheet
        product_workbook = load_workbook(product_file)
        product_sheet = product_workbook.active

        # Load the branch file and its sheets
        branch_workbook = load_workbook(branch_file)
        sales_sheet = branch_workbook["Daily Sales in branch"]
        quantity_sheet = branch_workbook["Product Quantity"]

        # Copy data from the product file to the respective sheets in the branch file
        for row in product_sheet.iter_rows(min_row=2, max_row=product_sheet.max_row, max_col=3, values_only=True):
            sales_sheet.append(row)  # Append to "Sales in branch" sheet
            quantity_sheet.append(row)  # Append to "Product Quantity" sheet

        # Save the updated branch workbook
        branch_workbook.save(branch_file)

        print(f"Product data successfully copied to sheets in '{branch_file}'.")

# class price analysis- SRP
class Price_Analysis:
    @staticmethod
    def Price_change():
        try:
            # Load the Products file
            workbook = load_workbook(product_file)
            sheet = workbook.active
            # Convert the sheet into a pandas DataFrame
            data = sheet.values
            columns = next(data)
            price_data_frame = pd.DataFrame(data, columns=columns)
            # Set the product ID and product name as the index for easier filtering
            price_data_frame.set_index(['Product ID', 'Product'], inplace=True)
            # Extract the date columns dynamically (convert column names to strings)
            date_columns = [str(col) for col in price_data_frame.columns if '/' in str(col)]
            # Ask the user for the year to filter
            year = input("Enter the year for which you want to see price changes (e.g., 2023): ").strip()
            if not year.isdigit() or len(year) != 4:
                print("Invalid year format. Please enter a valid 4-digit year.")
                return
            # Filter the date columns by the selected year
            filtered_columns = [col for col in date_columns if col.endswith(year)]
            if not filtered_columns:
                print(f"No price data available for the year {year}.")
                return
            # Ask the user if they want to see all products or one product
            choice = input(
                "Do you want to see the price change for all products or a single product? (all/one): ").strip().lower()
            if choice not in ['all', 'one']:
                print("Invalid choice. Please enter 'all' or 'one'.")
                return
            if choice == "all":
                try:
                    # Plot price change for all products
                    if price_data_frame.empty:
                        print("No product data available to display.")
                        return
                    plt.figure(figsize=(12, 6))
                    for (product_id, product_name) in price_data_frame.index:
                        prices = price_data_frame.loc[(product_id, product_name), filtered_columns]
                        plt.plot(filtered_columns, prices, marker='o', label=f"{product_name} ({product_id})")
                    # Customize the graph
                    plt.title(f"Price Change for All Products in {year}")
                    plt.xlabel("Dates")
                    plt.ylabel("Prices")
                    plt.xticks(rotation=45)
                    plt.legend()
                    plt.tight_layout()
                    plt.show()
                except Exception as e:
                    print(f"An error occurred while plotting: {e}")
                    return
            elif choice == "one":
                # Ask for the Product ID or Product Name
                search_choice = input(
                    "Would you like to search by Product ID or Product Name? (id/name): ").strip().lower()
                if search_choice not in ['id', 'name']:
                    print("Invalid search choice. Please enter 'id' or 'name'.")
                    return
                product_rows = None
                if search_choice == "id":
                    product_id = input("Enter the Product ID: ").strip()
                    if product_id in price_data_frame.index.get_level_values('Product ID'):
                        product_rows = price_data_frame.loc[product_id]
                    else:
                        print(f"No product found with ID: {product_id}")
                        return
                elif search_choice == "name":
                    product_name = input("Enter the Product Name: ").strip()
                    if product_name in price_data_frame.index.get_level_values('Product'):
                        product_rows = price_data_frame.xs(product_name, level='Product')
                    else:
                        print(f"No product found with name: {product_name}")
                        return
                if product_rows is not None:
                    try:
                        # Handle single-row or multi-row cases
                        if isinstance(product_rows, pd.Series):  # Single row case
                            product_id, product_name = product_rows.name
                            prices = product_rows[filtered_columns]
                        else:  # Multi-row case
                            first_row = product_rows.iloc[0]
                            product_id = first_row.name[0]  # Extract Product ID
                            product_name = first_row.name[1]  # Extract Product Name
                            prices = first_row[filtered_columns]
                        # Plot price change for the specific product
                        plt.figure(figsize=(10, 5))
                        plt.plot(filtered_columns, prices, marker='o', label=product_name)
                        # Customize the graph
                        plt.title(f"Price Change for {product_name} ({product_id}) in {year}")
                        plt.xlabel("Dates")
                        plt.ylabel("Prices")
                        plt.xticks(rotation=45)
                        plt.legend()
                        plt.tight_layout()
                        plt.show()
                    except Exception as e:
                        print(f"An error occurred while plotting: {e}")
                        return
                else:
                    print("Product not found in the data.")
        except FileNotFoundError:
            print("The product file could not be found. Please check the file path.")
        except Exception as e:
            print(f"An unexpected error occurred: {e}")

# class sales
class Sales:
    @staticmethod
    def insert_branch_sales():
        if not Sampath_food_cities.branches:
            print("No branches available. Please insert a branch first.")
            return
        # Ask for Branch ID or Branch Name
        search_key = input("Enter the Branch ID or Branch Name: ")
        branch_id = None
        # Search the branch in the dictionary
        for b_id, b_info in Sampath_food_cities.branches.items():
            if search_key == b_id or search_key.lower() == b_info["name"].lower():
                branch_id = b_id
                break
        if not branch_id:
            print(f"No branch found with ID or Name '{search_key}'.")
            return
        # Get the consolidated branch file
        branch_file = f"{branch_id}.xlsx"
        if not os.path.exists(branch_file):
            print(f"Branch file for {branch_id} does not exist.")
            return
        # Load the branch file
        branch_workbook = load_workbook(branch_file)
        quantity_sheet = branch_workbook["Product Quantity"]  # Access the "Product Quantity" sheet
        # Ask for the date of the quantities
        date = input("Enter the date for quantities (DD/MM/YYYY): ")
        # Check if the date column already exists, otherwise add it
        date_col_index = None
        for col_index, cell in enumerate(quantity_sheet[1], start=1):
            if cell.value == date:
                date_col_index = col_index
                break
        if not date_col_index:
            date_col_index = quantity_sheet.max_column + 1
            quantity_sheet.cell(row=1, column=date_col_index, value=date)
        # Start inserting quantities for products
        print("Enter quantities for products. Type 'done' to finish.")
        while True:
            # Ask for Product ID or Product Name
            product_id_or_name = input("Enter Product ID or Name (or type 'done' to finish): ")
            if product_id_or_name.lower() == "done":
                break
            # Find the product in the sheet
            product_row_index = None
            for row_index, row in enumerate(quantity_sheet.iter_rows(min_row=2, max_row=quantity_sheet.max_row),
                                            start=2):
                if product_id_or_name == str(row[0].value) or product_id_or_name.lower() == str(row[1].value).lower():
                    product_row_index = row_index
                    break
            if not product_row_index:
                print(f"No product found with ID or Name '{product_id_or_name}'.")
                continue
            # Ask for the quantity
            try:
                quantity = int(
                    input(f"Enter quantity for {quantity_sheet.cell(row=product_row_index, column=2).value}: "))
            except ValueError:
                print("Invalid input for quantity. Please enter a valid number.")
                continue
            # Update the quantity in the correct cell
            quantity_sheet.cell(row=product_row_index, column=date_col_index, value=quantity)
        # Save the updated branch file
        branch_workbook.save(branch_file)
        print(f"Quantity details for branch '{Sampath_food_cities.branches[branch_id]['name']}' on {date} have been successfully updated.")
        Sales.update_branch_sales_from_quantity(branch_id, date)
        Monthly_Sales.update_sales_quantity_to_branch_monthly_sales_sheet(branch_id)

    @staticmethod
    def update_branch_sales_from_quantity(branch_id, date):
        # File name for the branch
        sales_file = f"{branch_id}.xlsx"
        # Check if the sales file exists
        if not os.path.exists(sales_file):
            print(f"Error: Sales file '{sales_file}' does not exist.")
            return
        # Load the sales workbook and access respective sheets
        sales_workbook = load_workbook(sales_file)
        quantity_sheet = sales_workbook["Product Quantity"]  # Access the "Product Quantity" sheet
        sales_sheet = sales_workbook["Daily Sales in branch"]  # Access the "Sales" sheet
        # Find the date column in the sales sheet
        sales_date_col_index = None
        for col_index, cell in enumerate(sales_sheet[1], start=1):
            if cell.value == date:
                sales_date_col_index = col_index
                break
        # If the date column doesn't exist, append a new column
        if not sales_date_col_index:
            print(f"Date '{date}' not found in the Sales sheet. Adding new column...")
            sales_date_col_index = len(sales_sheet[1]) + 1  # Set column index to the next available column
            sales_sheet.cell(row=1, column=sales_date_col_index, value=date)  # Add date as column header
        # Find the date column in the Product Quantity sheet
        quantity_date_col_index = None
        for col_index, cell in enumerate(quantity_sheet[1], start=1):
            if cell.value == date:
                quantity_date_col_index = col_index
                break
        if not quantity_date_col_index:
            print(f"Error: Date '{date}' does not exist in the Product Quantity sheet. Please insert the date first.")
            return
        # Initialize total sales for the date
        total_sales = 0
        # Loop through each product in the Product Quantity sheet
        for row_index, row in enumerate(quantity_sheet.iter_rows(min_row=2, max_row=quantity_sheet.max_row), start=2):
            product_id = row[0].value
            quantity = row[quantity_date_col_index - 1].value
            if not quantity:
                continue  # Skip if no quantity is entered
            # Find the product in the Sales sheet
            sales_row_index = None
            for s_row_index, s_row in enumerate(sales_sheet.iter_rows(min_row=3, max_row=sales_sheet.max_row), start=3):
                if s_row[0].value == product_id:
                    sales_row_index = s_row_index
                    break
            if not sales_row_index:
                print(f"Warning: Product ID '{product_id}' not found in Sales sheet. Skipping...")
                continue
            # Get the unit price
            unit_price = sales_sheet.cell(row=sales_row_index, column=3).value
            if not unit_price:
                print(f"Warning: Unit price not found for Product ID '{product_id}'. Skipping...")
                continue
            # Calculate the sales amount
            sales_amount = quantity * unit_price
            # Update the sales amount in the Sales sheet
            sales_sheet.cell(row=sales_row_index, column=sales_date_col_index, value=sales_amount)
            # Add to total sales
            total_sales += sales_amount
        # Update total sales in the second row under the date column
        sales_sheet.cell(row=2, column=sales_date_col_index, value=total_sales)
        # Save the updated sales workbook
        sales_workbook.save(sales_file)
        print(f"Sales details for branch '{branch_id}' on {date} have been successfully updated based on quantities.")
        print(f"Total sales for '{date}' in branch '{branch_id}': {total_sales}")
        # Additional updates (external methods)
        Sales.update_daily_network_sales(branch_id, date, total_sales)
        Weekly_Sales.update_weekly_sales()
        Monthly_Sales.update_monthly_sales()

    @staticmethod
    def update_daily_network_sales(branch_id, sales_date, total_sales):
        # Path to the network sales file
        network_sales_file = "Whole_Network_Sales.xlsx"
        # Create the file if it doesn't exist
        if not os.path.exists(network_sales_file):
            workbook = Workbook()
            sheet = workbook.active
            # Add headers
            sheet.cell(row=1, column=1, value="Branch ID")
            workbook.save(network_sales_file)
        # Load the network sales file
        network_workbook = load_workbook(network_sales_file)
        network_sheet = network_workbook.active
        # Check if the branch ID exists in the first column
        branch_row_index = None
        for row_index, cell in enumerate(network_sheet.iter_rows(min_col=1, max_col=1, min_row=2), start=2):
            if cell[0].value == branch_id:
                branch_row_index = row_index
                break
        if not branch_row_index:
            # Add new row for the branch
            branch_row_index = network_sheet.max_row + 1
            network_sheet.cell(row=branch_row_index, column=1, value=branch_id)
        # Check if the date column exists, otherwise add it
        date_col_index = None
        for col_index, cell in enumerate(network_sheet[1], start=1):
            if cell.value == sales_date:
                date_col_index = col_index
                break
        if not date_col_index:
            # Add new column for the date
            date_col_index = network_sheet.max_column + 1
            network_sheet.cell(row=1, column=date_col_index, value=sales_date)
        # Update the total sales for the branch under the date column
        network_sheet.cell(row=branch_row_index, column=date_col_index, value=total_sales)
        # Recalculate the total sales for the date across all branches
        total_sales_for_date = 0
        for row in network_sheet.iter_rows(min_row=3, max_row=network_sheet.max_row, min_col=date_col_index,
                                           max_col=date_col_index):
            total_sales_for_date += row[0].value or 0
        # Update the second row with the total sales for the date across all branches
        network_sheet.cell(row=2, column=date_col_index, value=total_sales_for_date)
        # Save the updated network sales file
        network_workbook.save(network_sales_file)
        print(f"Total sales for branch '{branch_id}' on {sales_date} have been updated in the network sales file.")

    @staticmethod
    def find_or_create_row(sheet, branch_id):
        for row_index, cell in enumerate(sheet.iter_rows(min_col=1, max_col=1, min_row=2), start=2):
            if cell[0].value == branch_id:
                return row_index

        # Create a new row if branch_id not found
        new_row_index = sheet.max_row + 1
        sheet.cell(row=new_row_index, column=1, value=branch_id)
        return new_row_index

    @staticmethod
    def find_or_create_column(sheet, header):
        for col_index, cell in enumerate(sheet[1], start=1):
            if cell.value == header:
                return col_index

        # Create a new column if header not found
        new_col_index = sheet.max_column + 1
        sheet.cell(row=1, column=new_col_index, value=header)
        return new_col_index

    @staticmethod
    def display_monthly_product_preference_of_branch():
        # Ask for Branch ID and locate the file
        search_key = input("Enter the Branch ID or Branch Name: ")
        # Ensure branches are available
        if not Sampath_food_cities.branches:
            print("No branches available. Please insert a branch first.")
            return
        branch_id = None
        # Search the branch in the dictionary
        for b_id, b_info in Sampath_food_cities.branches.items():
            if search_key == b_id or search_key.lower() == b_info["name"].lower():
                branch_id = b_id
                break
        if not branch_id:
            print(f"No branch found with ID or Name '{search_key}'.")
            return
        # Get the branch file
        file_name = f"{branch_id}.xlsx"
        try:
            # Open the workbook and Monthly Sales sheet
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook["Monthly Sales"]
        except FileNotFoundError:
            print(f"Error: File '{file_name}' not found.")
            return
        except KeyError:
            print("Error: 'Monthly Sales' sheet not found.")
            return
        # Ask for the month and year to filter columns
        month = input("Enter the Month (e.g., January): ").capitalize()
        year = input("Enter the Year (e.g., 2024): ")
        target_column = None
        # Find the target column for the given month and year
        for col in range(3, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header and f"{month} {year}" == header:
                target_column = col
                break
        if not target_column:
            print(f"No data found for {month} {year}.")
            return
        # Extract product names and their sold quantities
        product_data = []
        for row in range(2, sheet.max_row + 1):
            product_name = sheet.cell(row=row, column=2).value
            quantity_sold = sheet.cell(row=row, column=target_column).value
            if quantity_sold is not None:
                product_data.append((product_name, quantity_sold))
        if not product_data:
            print(f"No sales data available for {month} {year} in branch {branch_id}.")
            return
        # Display the data in a table format
        print("\nMonthly Sales Data for All Products")
        print(f"{'Product Name':<30}{'Quantity Sold':<15}")
        print("-" * 45)
        for product_name, quantity_sold in product_data:
            print(f"{product_name:<30}{quantity_sold:<15}")
        # Plot the bar chart
        product_names = [product[0] for product in product_data]
        quantities = [product[1] for product in product_data]
        plt.figure(figsize=(12, 6))
        plt.bar(product_names, quantities, color='orange')
        plt.title(f"Monthly Sales: {month} {year} (Branch {branch_id})", fontsize=16)
        plt.xlabel("Products", fontsize=14)
        plt.ylabel("Quantity Sold", fontsize=14)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        plt.show()

# class weekly sales
class Weekly_Sales:
    @staticmethod
    def update_weekly_sales():
        network_sales_file = "Whole_Network_Sales.xlsx"
        weekly_sales_file = "Whole_Network_Weekly_Sales.xlsx"
        # Create the weekly sales file if it doesn't exist
        if not os.path.exists(weekly_sales_file):
            network_files_creator.create_network_files(weekly_sales_file, "Weekly Network")
        # Load both files
        network_workbook = load_workbook(network_sales_file)
        network_sheet = network_workbook.active
        weekly_workbook = load_workbook(weekly_sales_file)
        weekly_sheet = weekly_workbook.active
        # Extract dates from the first row of the network sales file
        date_columns = [
            (col_index, cell.value) for col_index, cell in enumerate(network_sheet[1], start=1) if col_index > 1
        ]
        # Dictionary to track total sales for each week across all branches
        total_weekly_sales_summary = {}
        for branch_row in network_sheet.iter_rows(
                min_row=3, max_row=network_sheet.max_row, min_col=1, max_col=network_sheet.max_column
        ):
            branch_id = branch_row[0].value
            total_weekly_sales = {}
            for col_index, sales_date in date_columns:
                if not sales_date:
                    continue
                # Parse the date and calculate the week number and year
                date_obj = datetime.strptime(sales_date, "%d/%m/%Y")
                week_number = date_obj.isocalendar()[1]
                year = date_obj.year
                week_key = f"Week {week_number} ({year})"
                # Accumulate sales
                sales_amount = branch_row[col_index - 1].value or 0
                total_weekly_sales[week_key] = total_weekly_sales.get(week_key, 0) + sales_amount
                total_weekly_sales_summary[week_key] = total_weekly_sales_summary.get(week_key, 0) + sales_amount
            # Update the weekly sales file for the branch
            branch_row_index = Sales.find_or_create_row(weekly_sheet, branch_id)
            for week_key, weekly_total in total_weekly_sales.items():
                week_col_index = Sales.find_or_create_column(weekly_sheet, week_key)
                weekly_sheet.cell(row=branch_row_index, column=week_col_index, value=weekly_total)
        # Update the second row with total weekly sales across all branches
        for week_key, weekly_total in total_weekly_sales_summary.items():
            week_col_index = Sales.find_or_create_column(weekly_sheet, week_key)
            weekly_sheet.cell(row=2, column=week_col_index, value=weekly_total)
        # Save the updated weekly sales file
        weekly_workbook.save(weekly_sales_file)
        print("Weekly sales updated successfully.")

    @staticmethod
    def weekly_sales_analysis():
        # Path to the weekly sales file
        weekly_sales_file = "Whole_Network_Weekly_Sales.xlsx"
        # Get the month input from the user
        user_month = input("Enter the month for the weekly sales analysis (e.g., December): ").capitalize()
        user_year = int(input("Enter the year (e.g., 2024): "))
        # Load the weekly sales file
        weekly_workbook = load_workbook(weekly_sales_file)
        weekly_sheet = weekly_workbook.active
        # Extract weeks (from 1st row, starting from the 2nd column)
        weeks = [cell.value for cell in weekly_sheet[1][1:]]
        # Determine the starting date of each week
        def week_to_month(week_str):
            try:
                # Parse week and year from the string "Week XX (YYYY)"
                parts = week_str.replace("Week ", "").split(" ")
                week_number = int(parts[0])
                year = int(parts[1].strip("()"))
                # Calculate the first day of the year and derive the week's start date
                first_day_of_year = datetime(year, 1, 1)
                week_start_date = first_day_of_year + timedelta(days=(week_number - 1) * 7)
                return week_start_date.strftime("%B"), year
            except Exception as e:
                print(f"Error parsing week string '{week_str}': {e}")
                return None, None
        # Filter weeks that belong to the specified month and year
        filtered_weeks = [
            week for week in weeks
            if week_to_month(week)[0] == user_month and week_to_month(week)[1] == user_year
        ]
        # Handle case where no weeks match the user's input
        if not filtered_weeks:
            print(f"No data available for {user_month} {user_year}.")
            return
        # Get column indices of the filtered weeks
        week_indices = [weeks.index(week) + 2 for week in filtered_weeks]  # +2 because column indices are 1-based
        # Extract total sales data for the filtered weeks
        total_sales = []
        for col_index in week_indices:
            cell_value = weekly_sheet.cell(row=2, column=col_index).value
            if cell_value is not None:
                total_sales.append(cell_value)
            else:
                total_sales.append(0)  # Default to 0 if the cell is empty
        # Prepare branch-wise sales data for the filtered weeks
        branches = []
        branch_sales = []
        for row in weekly_sheet.iter_rows(min_row=3, max_row=weekly_sheet.max_row, min_col=1,
                                          max_col=weekly_sheet.max_column):
            branch_id = row[0].value
            if branch_id is not None:
                branches.append(branch_id)
                sales = []
                for col_index in week_indices:
                    cell_value = row[col_index - 1].value  # Adjust for 0-based index
                    sales.append(cell_value if cell_value is not None else 0)  # Default to 0 if empty
                branch_sales.append(sales)
        # Create a DataFrame for the table view
        data = {'Week': filtered_weeks}
        for i, branch in enumerate(branches):
            data[branch] = [branch_sales[i][j] for j in range(len(filtered_weeks))]
        data['Total Sales'] = total_sales
        # Convert to DataFrame
        df = pd.DataFrame(data)
        # Display the table
        print(f"\nWeekly Sales Data for {user_month} {user_year}:")
        print(df)
        # Plot total sales for the network (Total Sales row)
        plt.figure(figsize=(10, 6))
        plt.bar(filtered_weeks, total_sales, color='green', label='Total Sales', alpha=0.6)
        # Plot each branch's sales (optional)
        for i, branch in enumerate(branches):
            plt.plot(filtered_weeks, [branch_sales[i][j] for j in range(len(filtered_weeks))], marker='o', label=branch)
        # Customize the plot
        plt.xlabel('Weeks')
        plt.ylabel('Sales Amount')
        plt.title(f'Weekly Sales Analysis for {user_month} {user_year}')
        plt.xticks(rotation=45)  # Rotate the week labels for better readability
        plt.legend(title="Branches", loc="upper left")
        # Show the plot
        plt.tight_layout()
        plt.show()

# class monthly sales
class Monthly_Sales:
    @staticmethod
    def update_sales_quantity_to_branch_monthly_sales_sheet(branch_id):
        # Load the branch file
        branch_file = f"{branch_id}.xlsx"
        if not os.path.exists(branch_file):
            print(f"Branch file for {branch_id} does not exist.")
            return
        branch_workbook = load_workbook(branch_file)
        # Access the "Monthly Sales" sheet or create it if it doesn't exist
        if "Monthly Sales" not in branch_workbook.sheetnames:
            monthly_sales_sheet = branch_workbook.create_sheet("Monthly Sales")
            monthly_sales_sheet.append(["Product ID", "Product Name"])  # Add headers for Product ID and Product Name
        else:
            monthly_sales_sheet = branch_workbook["Monthly Sales"]
        # Access the "Monthly Purchases" sheet
        if "Monthly Purchases" not in branch_workbook.sheetnames:
            print("Error: 'Monthly Purchases' sheet not found in branch file.")
            return
        monthly_purchases_sheet = branch_workbook["Monthly Purchases"]
        # Access the "Product Quantity" sheet
        if "Product Quantity" not in branch_workbook.sheetnames:
            print("Error: 'Product Quantity' sheet not found in branch file.")
            return
        quantity_sheet = branch_workbook["Product Quantity"]
        # Create a dictionary to track total quantities per product per month
        monthly_totals = {}
        # Loop through the "Product Quantity" sheet to aggregate data by month
        for col in quantity_sheet.iter_cols(min_row=1, max_row=1, min_col=3, max_col=quantity_sheet.max_column):
            date_cell = col[0].value
            if not date_cell:
                continue
            # Extract month and year from the date
            try:
                date = datetime.strptime(date_cell, "%d/%m/%Y")
                month = date.strftime("%B %Y")  # e.g., "January 2024"
            except ValueError:
                # print(f"Invalid date format in 'Product Quantity': {date_cell}")
                continue
            # Aggregate the quantities for this month
            for row in quantity_sheet.iter_rows(min_row=2, max_row=quantity_sheet.max_row):
                product_id = row[0].value
                product_name = row[1].value
                quantity = row[col[0].column - 1].value or 0  # Get quantity from the corresponding date column
                if not product_id or not product_name:
                    continue
                # Initialize data for the product if not already done
                if product_id not in monthly_totals:
                    monthly_totals[product_id] = {"name": product_name, "totals": {}}
                # Add the quantity to the respective month
                if month not in monthly_totals[product_id]["totals"]:
                    monthly_totals[product_id]["totals"][month] = 0
                monthly_totals[product_id]["totals"][month] += quantity
        # Update the "Monthly Sales" sheet with aggregated data
        for product_id, product_data in monthly_totals.items():
            product_name = product_data["name"]
            for month, total_quantity in product_data["totals"].items():
                # Find or create the month column in the "Monthly Sales" sheet
                month_col_index = None
                for col_index, cell in enumerate(monthly_sales_sheet[1], start=1):
                    if cell.value == month:
                        month_col_index = col_index
                        break
                if not month_col_index:
                    month_col_index = monthly_sales_sheet.max_column + 1
                    monthly_sales_sheet.cell(row=1, column=month_col_index, value=month)
                # Find or create the product row
                product_row_index = None
                for row_index, row in enumerate(
                        monthly_sales_sheet.iter_rows(min_row=2, max_row=monthly_sales_sheet.max_row), start=2):
                    if row[0].value == product_id:
                        product_row_index = row_index
                        break
                if not product_row_index:
                    product_row_index = monthly_sales_sheet.max_row + 1
                    monthly_sales_sheet.cell(row=product_row_index, column=1, value=product_id)
                    monthly_sales_sheet.cell(row=product_row_index, column=2, value=product_name)
                # Check the purchase value from the "Monthly Purchases" sheet
                purchase_col_index = None
                for col_index, cell in enumerate(monthly_purchases_sheet[1], start=1):
                    if cell.value == month:
                        purchase_col_index = col_index
                        break
                if purchase_col_index:
                    purchase_value = None
                    for row in monthly_purchases_sheet.iter_rows(min_row=2, max_row=monthly_purchases_sheet.max_row):
                        if row[0].value == product_id:
                            purchase_value = row[purchase_col_index - 1].value or 0
                            break
                    if purchase_value is not None and total_quantity > purchase_value:
                        print(f"Warning: Sales for Product ID {product_id} ({product_name}) in {month} exceeds purchases.")
                        total_quantity = purchase_value
                # Update the total quantity for the product in the month's column
                monthly_sales_sheet.cell(row=product_row_index, column=month_col_index, value=total_quantity)
        # Save the updated branch file
        branch_workbook.save(branch_file)
        print(f"Monthly sales updated for branch '{branch_id}' with purchase validation.")

    @staticmethod
    def update_monthly_sales():
        network_sales_file = "Whole_Network_Sales.xlsx"
        monthly_sales_file = "Whole_Network_Monthly_Sales.xlsx"
        # Create the monthly sales file if it doesn't exist
        if not os.path.exists(monthly_sales_file):
            network_files_creator.create_network_files(monthly_sales_file, "Monthly Network")
        # Load both files
        network_workbook = load_workbook(network_sales_file)
        network_sheet = network_workbook.active
        monthly_workbook = load_workbook(monthly_sales_file)
        monthly_sheet = monthly_workbook.active
        # Extract dates from the first row of the network sales file
        date_columns = [(col_index, cell.value) for col_index, cell in enumerate(network_sheet[1], start=1) if
                        col_index > 1]
        # Dictionary to track total sales for each month across all branches
        total_monthly_sales_summary = {}
        for branch_row in network_sheet.iter_rows(min_row=3, max_row=network_sheet.max_row, min_col=1,
                                                  max_col=network_sheet.max_column):
            branch_id = branch_row[0].value
            total_monthly_sales = {}
            for col_index, sales_date in date_columns:
                if not sales_date:
                    continue
                # Format the month and year as "Month Year"
                month_year = datetime.strptime(sales_date, "%d/%m/%Y").strftime("%B %Y")
                sales_amount = branch_row[col_index - 1].value or 0
                total_monthly_sales[month_year] = total_monthly_sales.get(month_year, 0) + sales_amount
                total_monthly_sales_summary[month_year] = total_monthly_sales_summary.get(month_year, 0) + sales_amount
            # Update the monthly sales file for the branch
            branch_row_index = Sales.find_or_create_row(monthly_sheet, branch_id)
            for month_year, monthly_total in total_monthly_sales.items():
                month_col_index = Sales.find_or_create_column(monthly_sheet, month_year)
                monthly_sheet.cell(row=branch_row_index, column=month_col_index, value=monthly_total)
        # Update the second row with total monthly sales across all branches
        for month_year, monthly_total in total_monthly_sales_summary.items():
            month_col_index = Sales.find_or_create_column(monthly_sheet, month_year)
            monthly_sheet.cell(row=2, column=month_col_index, value=monthly_total)
        monthly_workbook.save(monthly_sales_file)
        print("Monthly sales updated successfully.")

    @staticmethod
    def monthly_sales_analysis():
        # Path to the network sales file
        network_sales_file = "Whole_Network_Monthly_Sales.xlsx"
        # Get the year input from the user
        user_year = input("Enter the year for the monthly sales analysis (e.g., 2024): ")
        # Load the network sales file
        network_workbook = load_workbook(network_sales_file)
        network_sheet = network_workbook.active
        # Extract months (from the second row, starting from the second column)
        months = [cell.value for cell in network_sheet[1][1:]]  # Assumes months start from the second column
        # Filter months that match the entered year (e.g., 'December 2024')
        filtered_months = [month for month in months if user_year in str(month)]
        # Ensure filtered_months has valid data before proceeding
        if not filtered_months:
            print(f"No data found for the year {user_year}. Please check your data.")
            return
        # Get column indices of the filtered months
        month_indices = [months.index(month) + 2 for month in filtered_months]  # +2 because columns are 1-based
        # Extract total sales data for the filtered months
        total_sales = [network_sheet.cell(row=2, column=col_index).value for col_index in month_indices]
        # Prepare branch-wise sales data for the filtered months
        branches = []
        branch_sales = []
        # Iterate through rows to collect branch sales data
        for row in network_sheet.iter_rows(min_row=3, max_row=network_sheet.max_row, min_col=1,
                                           max_col=network_sheet.max_column):
            branch_id = row[0].value
            branches.append(branch_id)
            sales = [row[col_index - 1].value for col_index in month_indices]  # Sales data for each branch
            branch_sales.append(sales)
        # Create a DataFrame for the table view
        data = {'Month': filtered_months}
        for i, branch in enumerate(branches):
            data[branch] = [branch_sales[i][j] for j in range(len(filtered_months))]
        # Add total sales as the first row
        data['Total Sales'] = total_sales
        # Convert to DataFrame
        df = pd.DataFrame(data)
        # Display the table
        print("\nMonthly Sales Data for year", user_year)
        print(df)
        # Plot total sales for the network (Total Sales row)
        plt.figure(figsize=(10, 6))
        plt.bar(filtered_months, total_sales, color='blue', label='Total Sales', alpha=0.6)
        # Plot each branch's sales (optional)
        for i, branch in enumerate(branches):
            plt.plot(filtered_months, [branch_sales[i][j] for j in range(len(filtered_months))], marker='o',
                     label=branch)
        # Customize the plot
        plt.xlabel('Months')
        plt.ylabel('Sales Amount')
        plt.title(f'Monthly Sales Analysis of Whole Network ({user_year})')
        plt.xticks(rotation=45)  # Rotate the month labels for better readability
        plt.legend(title="Branches", loc="upper left")
        # Show the plot
        plt.tight_layout()
        plt.show()

# Single responsive-class sales distribution
class Sales_Distribution:
    @staticmethod
    def monthly_product_sales_distribution():
        # Step 1: Collect input
        search_key = input("Enter the Branch ID or Branch Name: ")
        month = input("Enter Month (e.g., January): ").capitalize()
        year = input("Enter Year (YYYY): ")
        # Ensure branches are available
        if not Sampath_food_cities.branches:
            print("No branches available. Please insert a branch first.")
            return
        branch_id = None
        # Search the branch in the dictionary
        for b_id, b_info in Sampath_food_cities.branches.items():
            if search_key == b_id or search_key.lower() == b_info["name"].lower():
                branch_id = b_id
                break
        if not branch_id:
            print(f"No branch found with ID or Name '{search_key}'.")
            return
        # Get the branch file
        file = f"{branch_id}.xlsx"
        try:
            # Step 2: Load data from Excel sheets
            purchases_dataframe = pd.read_excel(file, sheet_name="Monthly Purchases")
            sales_dataframe = pd.read_excel(file, sheet_name="Monthly Sales")
            # Step 3: Reshape data for processing
            purchases_dataframe = purchases_dataframe.melt(id_vars=["Product ID", "Product Name"], var_name="Month",
                                             value_name="Purchases")
            sales_dataframe = sales_dataframe.melt(id_vars=["Product ID", "Product Name"], var_name="Month", value_name="Sales")
            # Merge purchases and sales data
            merged_dataframe = pd.merge(purchases_dataframe, sales_dataframe, on=["Product ID", "Product Name", "Month"])
            # Extract Year from Month column
            merged_dataframe["Year"] = merged_dataframe["Month"].apply(lambda x: int(x.split()[-1]))
            merged_dataframe["Month"] = merged_dataframe["Month"].apply(lambda x: x.split()[0])
            # Filter data for the specified year and month
            filtered_data = merged_dataframe[(merged_dataframe["Year"] == int(year)) & (merged_dataframe["Month"] == month)]
            if filtered_data.empty:
                print(f"No data available for {month} {year} in branch {branch_id}.")
                return
            # Calculate sales distribution for each product
            filtered_data["Sales Distribution (%)"] = (filtered_data["Sales"] / filtered_data["Purchases"]) * 100
            # Step 4: Display data as a table
            print("\nProduct-wise Sales Distribution Data")
            print(filtered_data[["Product Name", "Purchases", "Sales", "Sales Distribution (%)"]])
            # Step 5: Plot bar chart
            plt.figure(figsize=(12, 6))
            plt.bar(filtered_data["Product Name"], filtered_data["Sales Distribution (%)"], color='green')
            plt.xlabel("Product Name")
            plt.ylabel("Sales Distribution (%)")
            plt.title(f"Product-wise Sales Distribution for {month} {year} (Branch {search_key})")
            plt.xticks(rotation=45, ha='right')
            plt.ylim(0, 100)  # Assuming percentages won't exceed 100
            plt.grid(axis='y', linestyle='--', alpha=0.7)
            plt.tight_layout()
            plt.show()
        except FileNotFoundError:
            print(f"Error: File for branch {branch_id} not found.")
        except KeyError as e:
            print(f"Error: Missing expected column in data: {e}")
        except Exception as e:
            print(f"An error occurred: {e}")

# Singleton pattern: only one object can be there in the system.
class Menu:
    _instance = None  # Class attribute to hold the single instance

    def __new__(cls, branch_file):
        if cls._instance is None:
            cls._instance = super(Menu, cls).__new__(cls)
        return cls._instance

    def Main_Menu(self):
        print("1. Manage branches")
        print("2. Manage products")
        print("3. Monthly purchase")
        print("4. Branch sales entry section")
        print("5. View analysis section")
        print("6. Exit")
        choice = int(input("Enter your choice: "))
        if choice == 1:
            self.Branches_management_menu()
        elif choice == 2:
            self.Products_management_menu()
        elif choice == 3:
            self.Branch_purchase_management_menu()
        elif choice == 4:
            Sales.insert_branch_sales()
            self.Main_Menu()
        elif choice == 5:
            self.Analysis_Menu()
        elif choice == 6:
            exit()
        else:
            print("Invalid choice......")
            self.Main_Menu()

    # Have to complete implementation
    def Branches_management_menu(self):
        print("1. Add a new branch")
        print("2. Update branch")
        print("3. Delete branch")
        print("4. Search branch")
        print("5. Back to main menu")
        print("6. Exit")
        choice = int(input("Enter your choice: "))
        if choice == 1:
            Branch.insert()
            self.Branches_management_menu()
        elif choice == 2:
            Branch.update()
            self.Branches_management_menu()
        elif choice == 3:
            Branch.delete()
            self.Branches_management_menu()
        elif choice == 4:
            Branch.search()
            self.Branches_management_menu()
        elif choice == 5:
            self.Main_Menu()
        elif choice == 6:
            exit()
        else:
            print("Invalid choice......")
            self.Branches_management_menu()

    def Products_management_menu(self):
        print("1. Add a new product")
        print("2. Update product price")
        print("3. Delete product")
        print("4. Search product")
        print("5. Back to main menu")
        print("6. Exit")
        choice = int(input("Enter your choice: "))
        if choice == 1:
            Product.insert()
            self.Products_management_menu()
        elif choice == 2:
            Product.update()
            self.Products_management_menu()
        elif choice == 3:
            Product.delete()
            self.Products_management_menu()
        elif choice == 4:
            Product.search()
            self.Products_management_menu()
        elif choice == 5:
            self.Main_Menu()
        elif choice == 6:
            exit()
        else:
            print("Invalid choice......")
            self.Products_management_menu()

    def Branch_purchase_management_menu(self):
        print("1. Insert purchase for a branch")
        print("2. Update branch purchase")
        print("3. Delete purchase")
        print("4. Search purchase")
        print("5. Back to main menu")
        print("6. Exit")
        choice = int(input("Enter your choice: "))
        if choice == 1:
            Purchase.insert()
            self.Branch_purchase_management_menu()
        elif choice == 2:
            Purchase.update()
            self.Branch_purchase_management_menu()
        elif choice == 3:
            Purchase.delete()
            self.Branch_purchase_management_menu()
        elif choice == 4:
            Purchase.search()
            self.Branch_purchase_management_menu()
        elif choice == 5:
            self.Main_Menu()
        elif choice == 6:
            exit()
        else:
            print("Invalid choice......")
            self.Branch_purchase_management_menu()

    def Analysis_Menu(self):
        print("1. Monthly sales analysis ")
        print("2. Price analysis of each product")
        print("3. Weekly sales analysis of supermarket network")
        print("4. Product preference analysis")
        print("5. Branches' distribution of total sales amount of purchases")
        print("6. Back to main menu")
        print("7. Exit")
        choice_main_menu = int(input("Enter you choice: "))
        if choice_main_menu == 1:
            Monthly_Sales.monthly_sales_analysis()
            self.Analysis_Menu()
        elif choice_main_menu == 2:
            Price_Analysis.Price_change()
            self.Analysis_Menu()
        elif choice_main_menu == 3:
            Weekly_Sales.weekly_sales_analysis()
            self.Analysis_Menu()
        elif choice_main_menu == 4:
            Sales.display_monthly_product_preference_of_branch()
            self.Analysis_Menu()
        elif choice_main_menu == 5:
            Sales_Distribution.monthly_product_sales_distribution()
            self.Analysis_Menu()
        elif choice_main_menu == 6:
            self.Main_Menu()
        elif choice_main_menu == 7:
            exit()
        else:
            print("Invalid choice......")
            self.Analysis_Menu()

class Base_network_files_creator:
    def create_files(self):  # This is the template method
        self.create_daily_sales()
        self.create_weekly_sales()
        self.create_monthly_sales()

    @staticmethod
    def create_network_files(file_name, sheet_title):
        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.title = sheet_title
        work_sheet.append(["Branch ID"])
        work_sheet.cell(row=2, column=1, value="Total Sales")
        work_book.save(file_name)
        print(f"Sales file {file_name} created successfully.")

    # These would be abstract methods in a true template pattern
    def create_daily_sales(self):
        pass

    def create_weekly_sales(self):
        pass

    def create_monthly_sales(self):
        pass

class network_files_creator(Base_network_files_creator):
    def create_daily_sales(self):
        self.create_network_files("Whole_Network_Sales.xlsx", "Network")

    def create_weekly_sales(self):
        self.create_network_files("Whole_Network_Weekly_Sales.xlsx", "Weekly Network")

    def create_monthly_sales(self):
        self.create_network_files("Whole_Network_Monthly_Sales.xlsx", "Monthly Network")

class Sampath_food_cities():
    branches = {}  # Class variable to manage branches by their ID

    @classmethod
    def load_branches_from_excel(cls, file_name):
        if os.path.exists(file_name):
            work_book = load_workbook(file_name)
            work_sheet = work_book.active

            # Clear dictionary and reload from the sheet
            cls.branches.clear()
            for row in work_sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                branch_id, branch_name, location, contact_number = row
                cls.branches[branch_id] = {
                    "name": branch_name,
                    "location": location,
                    "contact_number": contact_number,
                    "file": f"{branch_id}.xlsx"
                }
            print("Branches loaded successfully.")
        else:
            print("Branch file not found. Starting with an empty dictionary.")

    @staticmethod
    def create_branches_file():
        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.title = "Branches"
        work_sheet.append(["Branch ID", "Branch Name", "Location", "Contact number"])
        work_book.save(branch_file)

        # Object creation for the network_files_creator class
        network_files=network_files_creator()
        network_files.create_files()

    @staticmethod
    def create_branch_management_file(branch_id):
        # Create a single workbook for the branch
        branch_file = f"{branch_id}.xlsx"
        branch_workbook = Workbook()

        # Create and format the "Sales in branch" sheet
        sales_worksheet = branch_workbook.active
        sales_worksheet.title = "Daily Sales in branch"
        sales_worksheet.append(["Product ID", "Product", "Unit Price"])
        sales_worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
        sales_worksheet.cell(row=2, column=1, value="Total Sales")

        # Create and format the "Product Quantity" sheet
        quantity_worksheet = branch_workbook.create_sheet(title="Product Quantity")
        quantity_worksheet.append(["Product ID", "Product", "Unit Price"])

        # Create and format the "Monthly Purchases" sheet
        purchase_worksheet = branch_workbook.create_sheet(title="Monthly Purchases")
        purchase_worksheet.append(["Product ID", "Product Name"])

        # Create and format the "Monthly Sales" sheet
        monthly_sales_worksheet = branch_workbook.create_sheet(title="Monthly Sales")
        monthly_sales_worksheet.append(["Product ID", "Product Name"])

        # Save the workbook
        branch_workbook.save(branch_file)
        print(f"Branch file '{branch_file}' with multiple sheets created successfully.")
        Update_Product_info.copy_products_to_branch_management_file(branch_id)

    @staticmethod
    def create_Products_file():
        # Create a new product file with headers
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Products"
        worksheet.append(["Product ID", "Product", "Unit Price"])
        workbook.save("Products.xlsx")

if __name__ == "__main__":
    # saving branches file name in the variable.
    branch_file = "Branches.xlsx"
    # loading all the branches to the branches dictionary.
    Sampath_food_cities.load_branches_from_excel(branch_file)
    # saving products file name in the variable.
    product_file="Products.xlsx"
    menu = Menu(branch_file)
    menu.Main_Menu()